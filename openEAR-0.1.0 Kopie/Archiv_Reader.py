from openpyxl import load_workbook

def Read_Excel_File(Directory):
    
    wb = load_workbook(Directory)
    ws = wb.get_sheet_by_name("Sheet")

    
    list_names = ["Archive_dt_string", 
        "Archive_Session_Name", 
        "Archive_Data_Time", 
        "Archive_Data_Arousal", 
        "Archive_Data_Valence", 
        "Archive_Data_EmodbEmotionAnger", 
        "Archive_Data_EmodbEmotionBoredom", 
        "Archive_Data_EmodbEmotionDisgust", 
        "Archive_Data_EmodbEmotionFear", 
        "Archive_Data_EmodbEmotionHappiness", 
        "Archive_Data_EmodbEmotionNeutral", 
        "Archive_Data_EmodbEmotionSadness", 
        "Archive_Data_AbcAffectAgressiv", 
        "Archive_Data_AbcAffectCheerfull", 
        "Archive_Data_AbcAffectIntoxicated", 
        "Archive_Data_AbcAffectNervous", 
        "Archive_Data_AbcAffectNeutral", 
        "Archive_Data_AbcAffectTired", 
        "Archive_Data_Loi1", 
        "Archive_Data_Loi2", 
        "Archive_Data_Loi3"]

    # Create an empty dictionary to store the rows
    data = {}

    # Iterate over the rows and append each one to the data dictionary
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 21:
            break
        if i < 2:
            data[list_names[i]] = row[0]
        elif i>=2:
            data[list_names[i]] = [float(cell) for cell in row]
    
    

    return data

def Get_Data(Directory):
    return Read_Excel_File(Directory)




    

















































