from typing import Union, Callable
import customtkinter
import openpyxl
from Frames.Observer import *
import time
from Frames.GraphFrames import *
import threading


## Objekt, das die Start-Einstellungen speichert, die im New_Analysis_Window festgelegt werden ##

class Startupsettings(object):
    selected_audio_device = ""
    session_name = ""
    working_mode= ""


    # Welche Skala angezeigt werden soll, 1 = Anzeigen
    loi_scale = 1
    valence_scale = 1
    arousal_scale = 0

    #Ob Emo+Emoji oder Dual Emotions, 1 = Dual
    show_dual_emotions = False

    #Welches Analyse Modell soll für Graphen in Big Window genommen werden, 0 = EmoDB
    show_abc_graphs = "AbcAffect"


class GlobalStartStop():
    analysis_loop = False

class WindowDisplayed():
    hello = False
    big_analysis = False
    settings = False
    archive = False

class Emotions(object):
    
    ##emodb
    Anger = -1
    Boredom = 0
    Disgust = -1
    Fear = -1
    Happiness = 1
    Neutral = 0
    Sadness = -1






## Globale Gewichte, müssen nicht instanziert werden. Alle Klassen sollten sich hier die Gewichte rausziehen können
## TODO: Schnittstelle zu TestDataExtractor2

class Weights(Subject):
    #EmoDB Gewichte
    w_emodb_anger = 0
    w_emodb_boredom = 0
    w_emodb_disgust = 0
    w_emodb_fear = 0
    w_emodb_happiness = 0
    w_emodb_neutral = 0
    w_emodb_sadness = 0

    #observer = Observer()

    '''w_emodb_anger.attach(observer)
    w_emodb_boredom.attach(observer)
    w_emodb_disgust.attach(observer)
    w_emodb_fear.attach(observer)
    w_emodb_happiness.attach(observer)
    w_emodb_neutral.attach(observer)
    w_emodb_sadness.attach(observer)
'''
    #abcAffect Gewichte
    w_abc_agressiv = 0
    w_abc_cheerful = 0
    w_abc_intoxicated = 0
    w_abc_nervous = 0
    w_abc_neutral = 0
    w_abc_tired = 0
    '''#avic Interest Gewichte
    w_avic_loi1 = 0
    w_avic_loi2 = 0
    w_avic_loi3 = 0

    #Arousal Gewicht
    w_arousal = 0

    #Valence Gewicht
    w_valence = 0'''

    #Working Mode
    working_mode = ""

    ## Liest Gewichte aus Excel aus und schreibt diese auf die jeweils passenden Variablen, 
    ## die aus der ersten Spalte gezogen werden

    def Read_Weights_from_Excel(cls, selected_working_mode):
        wb = openpyxl.load_workbook('openEAR-0.1.0 Kopie/GUI/Default Weights/default_weights.xlsx')
        
        # Select the sheet
        ws = wb['Gewichte']

        #Select the right Column based on the selected working mode
        column_selected = 0
        if selected_working_mode == "Pitch – Session":
            column_selected = 1
        elif selected_working_mode == "Gespräch":
            column_selected = 2
        else:
            column_selected = 3
            print("Benutzerdefinierte Gewichte")

                

        # Create an empty list to store the numbers
        numbers = []

        #Lies die Variablennamen aus der Excel Spalte 1 aus
        weights_variables = []
        for row in ws.iter_rows(3,20, values_only = True):
            weights_variables.append(row[0])


        #Lies die Gewichte aus Excel Spalte x aus
        for row in ws.iter_rows(3, 20, values_only=True):
            # Append the number in the selected column to the list
            numbers.append(row[int(column_selected)])


        # Zahlen den zugehörigen Variablen zuordnen
        #if column_selected != 3:
        for name, value in zip(weights_variables, numbers):
            setattr(cls, name, value)
    
        '''for name, value in zip(weights_variables, numbers):
            exec(name + ".value = " + str(value))
        '''
       
        #print(weights_variables)
        print(numbers)
        print(weights_variables)
        print("Agressiv:    " + str(Weights.w_abc_agressiv))

    def set_working_mode(working_mode = str):
        Weights.working_mode = working_mode
        #Weights.Read_Weights_from_Excel()

    
#Weights.Read_Weights_from_Excel(Weights, "Gespräch")

class CustomSpinBox(customtkinter.CTkFrame):
    def __init__(self, *args,
                 width: int = 100,
                 height: int = 32,
                 step_size: Union[int, float] = 1,
                 command: Callable = None,
                 title = "",
                 **kwargs):
        super().__init__(*args, width=width, height=height, **kwargs)

        #self.weights = Weights()

        self.step_size = 1
        self.command = command

        self.title = customtkinter.CTkLabel(self, text = title, fg_color="transparent")
        self.title.grid(row = 0, column = 1)

        self.configure(fg_color=("gray78", "gray28"))  # set frame color

        self.grid_columnconfigure((0, 2), weight=0)  # buttons don't expand
        self.grid_columnconfigure(1, weight=0)  # entry expands

        self.entry = customtkinter.CTkEntry(self, width=width-(2*height), height=height-6, border_width=0)
        self.entry.grid(row=1, column=1, columnspan=1, padx=3, pady=3, sticky = "ew")

        self.subtract_button = customtkinter.CTkButton(self, text="-", width=height-6, height=height-6,
                                                       command=lambda:self.subtract_button_callback(title=title), fg_color="#1f538d")
        self.subtract_button.grid(row=1, column=0, padx=(10, 0), pady=3)


        self.add_button = customtkinter.CTkButton(self, text="+", width=height-6, height=height-6,
                                                  command=lambda:self.add_button_callback(title=title), fg_color="#1f538d")
        self.add_button.grid(row=1, column=2, padx=(0, 10), pady=3)

        # default value
        self.entry.insert(0, "0")


    def add_button_callback(self, title):
        self.title = title
        
        if self.command is not None:
            self.command()
        try:
            value = int(self.entry.get()) + self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
            #self.change_weight_command(title=self.title, new_value = value)
            print(self.title + ":   " + str(self.change_weight_command(title=self.title, new_value=value)))
        except ValueError:
            return print ("error")

    def subtract_button_callback(self, title):
        self.title = title
        if self.entry.get() != 0:
            if self.command is not None:
                self.command()
            try:
                current_value = int(self.entry.get())
                if current_value > 0:
                    value = current_value - self.step_size
                    self.entry.delete(0, "end")
                    self.entry.insert(0, value)
                    #self.change_weight_command(title=self.title, new_value=value)
                    print(self.title + ":   " + str(self.change_weight_command(title=self.title, new_value=value)))
            except ValueError:
                return print("error")

    def get(self) -> Union[float, None]:
        try:
            return int(self.entry.get())
        except ValueError:
            return None

    def set(self, value: VariableMonitor):
        self.entry.delete(0, "end")
        self.entry.insert(0, str(value))
    
    
    def change_weight_command_old(self, title, new_value):
        self.new_value = new_value
        self.title = title
        
        match self.title:
            
            ## emoDB Emotionen
            case "Anger" : 
                Weights.w_emodb_anger.value = self.new_value
                return Weights.w_emodb_anger
            case "Boredom" : 
                Weights.w_emodb_boredom.value = self.new_value
                return Weights.w_emodb_boredom.value
            case "Disgust" : 
                Weights.w_emodb_disgust.value = self.new_value
                return Weights.w_emodb_disgust
            case "Fear" : 
                Weights.w_emodb_fear.value = self.new_value
                return Weights.w_emodb_fear
            case "Happiness" : 
                Weights.w_emodb_happiness.value = self.new_value
                return Weights.w_emodb_happiness
            case "Neutral" : 
                Weights.w_emodb_neutral.value = self.new_value
                return Weights.w_emodb_neutral
            case "Sadness" : 
                Weights.w_emodb_sadness.value = self.new_value
                return Weights.w_emodb_sadness

            ## abcAffect Emotionen
            case "Agressiv" : 
                Weights.w_abc_agressiv = self.new_value
                return Weights.w_abc_agressiv
            case "Cheerful" : 
                Weights.w_abc_cheerful = self.new_value
                return Weights.w_abc_cheerful
            case "Intoxicated" : 
                Weights.w_abc_intoxicated = self.new_value
                return Weights.w_abc_intoxicated
            case "Nervous" : 
                Weights.w_abc_nervous = self.new_value
                return Weights.w_abc_nervous
            case "Neutral_abc_Affect" : 
                Weights.w_abc_neutral = self.new_value
                return Weights.w_abc_neutral
            case "Tired" : 
                Weights.w_abc_tired = self.new_value
                return Weights.w_abc_tired
            

            ## Default-Case
            case _:
                print("Ich konnte kein passendes Gewicht zu dem Titel dieses Widgets finden:   " + self.title)



    def change_weight_command(self, title, new_value):
            self.new_value = new_value
            self.title = title
            
            match self.title:
                
                ## emoDB Emotionen
                case "Anger" : 
                    Weights.w_emodb_anger = self.new_value
                    return Weights.w_emodb_anger
                case "Boredom" : 
                    Weights.w_emodb_boredom = self.new_value
                    return Weights.w_emodb_boredom
                case "Disgust" : 
                    Weights.w_emodb_disgust = self.new_value
                    return Weights.w_emodb_disgust
                case "Fear" : 
                    Weights.w_emodb_fear = self.new_value
                    return Weights.w_emodb_fear
                case "Happiness" : 
                    Weights.w_emodb_happiness = self.new_value
                    return Weights.w_emodb_happiness
                case "Neutral" : 
                    Weights.w_emodb_neutral = self.new_value
                    return Weights.w_emodb_neutral
                case "Sadness" : 
                    Weights.w_emodb_sadness = self.new_value
                    return Weights.w_emodb_sadness

                ## abcAffect Emotionen
                case "Agressiv" : 
                    Weights.w_abc_agressiv = self.new_value
                    print("geht")
                    return Weights.w_abc_agressiv
                case "Cheerful" : 
                    Weights.w_abc_cheerful = self.new_value
                    return Weights.w_abc_cheerful
                case "Intoxicated" : 
                    Weights.w_abc_intoxicated = self.new_value
                    return Weights.w_abc_intoxicated
                case "Nervous" : 
                    Weights.w_abc_nervous = self.new_value
                    return Weights.w_abc_nervous
                case "Neutral_abc_Affect" : 
                    Weights.w_abc_neutral = self.new_value
                    return Weights.w_abc_neutral
                case "Tired" : 
                    Weights.w_abc_tired = self.new_value
                    return Weights.w_abc_tired
                

                ## Default-Case
                case _:
                    print("Ich konnte kein passendes Gewicht zu dem Titel dieses Widgets finden:   " + self.title)




class Stopwatch(customtkinter.CTkFrame):
    

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.configure(fg_color = '#282928')
        
        
        self.time = 0.0
        self.running = False
        self.time_string = customtkinter.StringVar()
        self.time_string.set("There is no active Analysis.")
        self.current_time = customtkinter.CTkLabel(self, textvariable=self.time_string, font=customtkinter.CTkFont(size=20))
        self.current_time.grid(row = 0, column = 0, sticky = "nsew")

        #self.start()
        
    
    def start(self):
        self.running = True
        self.start_time = time.time()
        t3 = threading.Thread(target=self.update_time)
        t3.start()
        
        #self.update_time()

    
    def stop(self):
        self.running = False
        try:
            self.time = time.time() - self.start_time
        except AttributeError:
            self.time = time.time()
        minutes, seconds = divmod(self.time, 60)
        hours, minutes = divmod(minutes, 60)
        self.time_string.set("Total Analysis Duration:  %d:%02d:%02d" % (hours, minutes, seconds))

    def reset(self):
        self.running = False
        self.time = 0.0
        self.time_string.set("There is no active Analysis.")

   
    def update_time(self):
        while True and self.running:
            self.time = time.time() - self.start_time
            minutes, seconds = divmod(self.time, 60)
            hours, minutes = divmod(minutes, 60)
            self.time_string.set("Duration:  %d:%02d:%02d" % (hours, minutes, seconds))
            #self.current_time.after(10, self.update_time)
            time.sleep(0.1)