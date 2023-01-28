from openpyxl import load_workbook

def Read_Excel_File(Directory):
    
    wb = load_workbook(Directory)
    ws = wb.worksheets[0]



    # Create a list of list names
    list_names = [
        "Archive_Data_Time",
        "Archive_Data_Arousal",
        "Archive_Data_Valence",
        "Archive_Data_EmodbEmotionAnger",
        "Archive_Data_EmodbEmotionBoredom",
        "Archive_Data_EmodbEmotionDisgust",
        "Archive_Data_EmodbEmotionFear",
        "Archive_Data_EmodbEmotionHappiness",
        "Archive_Data_EmodbEmotionNeutral",
        "Archive_Data_EmodbEmotionSadness",
        "Archive_Data_AbcAffectAgressiv",
        "Archive_Data_AbcAffectCheerfull",
        "Archive_Data_AbcAffectIntoxicated",
        "Archive_Data_AbcAffectNervous",
        "Archive_Data_AbcAffectNeutral",
        "Archive_Data_AbcAffectTired",
        "Archive_Data_Loi1",
        "Archive_Data_Loi2",
        "Archive_Data_Loi3",
        'Archive_Soll_DataEmodbEmotionAnger',
        'Archive_Soll_DataEmodbEmotionBoredom',
        'Archive_Soll_DataEmodbEmotionDisgust',
        'Archive_Soll_DataEmodbEmotionFear',
        'Archive_Soll_DataEmodbEmotionHappiness',
        'Archive_Soll_DataEmodbEmotionNeutral',
        'Archive_Soll_DataEmodbEmotionSadness',
        'Archive_Soll_DataAbcAffectAgressiv',
        'Archive_Soll_DataAbcAffectCheerfull',
        'Archive_Soll_DataAbcAffectIntoxicated',
        'Archive_Soll_DataAbcAffectNervous',
        'Archive_Soll_DataAbcAffectNeutral',
        'Archive_Soll_DataAbcAffectTired',
        'Archive_Abs_MW_Data_Arousal',
        'Archive_Abs_MW_Data_Valence',
        'Archive_Abs_MW_Data_EmodbEmotionAnger',
        'Archive_Abs_MW_Data_EmodbEmotionBoredom',
        'Archive_Abs_MW_Data_EmodbEmotionDisgust',
        'Archive_Abs_MW_Data_EmodbEmotionFear',
        'Archive_Abs_MW_Data_EmodbEmotionHappiness',
        'Archive_Abs_MW_Data_EmodbEmotionNeutral',
        'Archive_Abs_MW_Data_EmodbEmotionSadness',
        'Archive_Abs_MW_Data_AbcAffectAgressiv',
        'Archive_Abs_MW_Data_AbcAffectCheerfull',
        'Archive_Abs_MW_Data_AbcAffectIntoxicated',
        'Archive_Abs_MW_Data_AbcAffectNervous',
        'Archive_Abs_MW_Data_AbcAffectNeutral',
        'Archive_Abs_MW_Data_AbcAffectTired',
        'Archive_Abs_MW_Data_Loi1',
        'Archive_Abs_MW_Data_Loi2',
        'Archive_Abs_MW_Data_Loi3',
        'Archive_Score_EmodbEmotions',
        'Archive_Score_AbcAffect',
        'Archive_Score_Retiva',
        'Archive_Abs_MW_Loi_Score'
    ]

    # Create an empty dictionary to store the rows
    data = {}

    # Iterate over the rows and append each one to the data dictionary
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 54:
            break
        data[list_names[i]] = [cell for cell in row if cell is not None]
    
    

    return data

def Get_Data(Directory):
    return Read_Excel_File(Directory)







#Beispiel Aufruf:

"""
data = Get_Data("/Users/paul/Documents/GitHub/ReTiVa3.0/openEAR-0.1.0 Kopie/SmileArchiv/test23_15:23_13_01_2023.xlsx")
Archive_Data_DateTime                           = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_DateTime", "Key not found")
Archive_Data_SessionNme                         = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_SessionName", "Key not found")
Archive_Data_Time                               = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_Time", "Key not found")
Archive_Data_Aroual                             = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_Arousal", "Key not found")
Archive_Data_Valence                            = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_Valence", "Key not found")
Archive_Data_EmodbEmtionAnger                   = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_EmodbEmotionAnger", "Key not found")
Archive_Data_EmodbEmotionBoredm                 = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_EmodbEmotionBoredom", "Key not found")
Archive_Data_EmodbEmotionDisgust                = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_EmodbEmotionDisgust", "Key not found")
Archive_Data_EmodbEmotionFear                   = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_EmodbEmotionFear", "Key not found")
Archive_Data_EmodbEmotionHappness               = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_EmodbEmotionHappiness", "Key not found")
Archive_Data_EmodbEmotionNeutral                = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_EmodbEmotionNeutral", "Key not found")
Archive_Data_EmodbEmotionSadness                = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_EmodbEmotionSadness", "Key not found")
Archive_Data_AbcAffectAgressiv                  = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_AbcAffectAgressiv", "Key not found")
Archive_Data_AbcAffectCheerful                  = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_AbcAffectCheerfull", "Key not found")
Archive_Data_AbcAffectIntoxicatd                = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_AbcAffectIntoxicated", "Key not found")
Archive_Data_AbcAffectNervous                   = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_AbcAffectNervous", "Key not found")
Archive_Data_AbcAffectNeutral                   = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_AbcAffectNeutral", "Key not found")
Archive_Data_AbcAffectTired                     = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_AbcAffectTired", "Key not found")
Archive_Data_Loi1                               = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_Loi1", "Key not found")
Archive_Data_Loi2                               = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_Loi2", "Key not found")
Archive_Data_Loi3                               = (Hier Müsste importierte Instanz stehen).data.get("Archive_Data_Loi3", "Key not found")
Archive_Soll_DataEmodbEmotionAnger               = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataEmodbEmotionAnger', "Key not found")
Archive_Soll_DataEmodbEmotionBoredom             = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataEmodbEmotionBoredom', "Key not found")
Archive_Soll_DataEmodbEmotionDisgust             = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataEmodbEmotionDisgust', "Key not found")
Archive_Soll_DataEmodbEmotionFear                = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataEmodbEmotionFear', "Key not found")
Archive_Soll_DataEmodbEmotionHappiness           = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataEmodbEmotionHappiness', "Key not found")
Archive_Soll_DataEmodbEmotionNeutral             = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataEmodbEmotionNeutral', "Key not found")
Archive_Soll_DataEmodbEmotionSadness             = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataEmodbEmotionSadness', "Key not found")
Archive_Soll_DataAbcAffectAgressiv               = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataAbcAffectAgressiv', "Key not found")
Archive_Soll_DataAbcAffectCheerfull              = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataAbcAffectCheerfull', "Key not found")
Archive_Soll_DataAbcAffectIntoxicated            = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataAbcAffectIntoxicated', "Key not found")
Archive_Soll_DataAbcAffectNervous                = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataAbcAffectNervous', "Key not found")
Archive_Soll_DataAbcAffectNeutral                = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataAbcAffectNeutral', "Key not found")
Archive_Soll_DataAbcAffectTired                  = (Hier Müsste importierte Instanz stehen).data.get('Archive_Soll_DataAbcAffectTired', "Key not found")
Archive_Abs_MW_Data_Arousal                      = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_Arousal', "Key not found")
Archive_Abs_MW_Data_Valence                      = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_Valence', "Key not found")
Archive_Abs_MW_Data_EmodbEmotionAnger            = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_EmodbEmotionAnger', "Key not found")
Archive_Abs_MW_Data_EmodbEmotionBoredom          = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_EmodbEmotionBoredom', "Key not found")
Archive_Abs_MW_Data_EmodbEmotionDisgust          = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_EmodbEmotionDisgust', "Key not found")
Archive_Abs_MW_Data_EmodbEmotionFear             = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_EmodbEmotionFear', "Key not found")
Archive_Abs_MW_Data_EmodbEmotionHappiness        = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_EmodbEmotionHappiness', "Key not found")
Archive_Abs_MW_Data_EmodbEmotionNeutral          = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_EmodbEmotionNeutral', "Key not found")
Archive_Abs_MW_Data_EmodbEmotionSadness          = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_EmodbEmotionSadness', "Key not found")
Archive_Abs_MW_Data_AbcAffectAgressiv            = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_AbcAffectAgressiv', "Key not found")
Archive_Abs_MW_Data_AbcAffectCheerfull           = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_AbcAffectCheerfull', "Key not found")
Archive_Abs_MW_Data_AbcAffectIntoxicated         = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_AbcAffectIntoxicated', "Key not found")
Archive_Abs_MW_Data_AbcAffectNervous             = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_AbcAffectNervous', "Key not found")
Archive_Abs_MW_Data_AbcAffectNeutral             = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_AbcAffectNeutral', "Key not found")
Archive_Abs_MW_Data_AbcAffectTired               = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_AbcAffectTired', "Key not found")
Archive_Abs_MW_Data_Loi1                         = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_Loi1', "Key not found")
Archive_Abs_MW_Data_Loi2                         = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_Loi2', "Key not found")
Archive_Abs_MW_Data_Loi3                         = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Data_Loi3', "Key not found")
Archive_Score_EmodbEmotions                      = (Hier Müsste importierte Instanz stehen).data.get('Archive_Score_EmodbEmotions', "Key not found")
Archive_Score_AbcAffect                          = (Hier Müsste importierte Instanz stehen).data.get('Archive_Score_AbcAffect', "Key not found")
Archive_Score_Retiva                             = (Hier Müsste importierte Instanz stehen).data.get('Archive_Score_Retiva', "Key not found")
Archive_Abs_MW_Loi_Score                         = (Hier Müsste importierte Instanz stehen).data.get('Archive_Abs_MW_Loi_Score', "Key not found")



print (Archive_Data_DateTime)
print (Archive_Data_SessionNme)
print (Archive_Data_Time)
print (Archive_Data_Aroual)
print (Archive_Data_Valence)
print (Archive_Data_EmodbEmtionAnger)
print (Archive_Data_EmodbEmotionBoredm)
print (Archive_Data_EmodbEmotionDisgust)
print (Archive_Data_EmodbEmotionFear)
print (Archive_Data_EmodbEmotionHappness)
print (Archive_Data_EmodbEmotionNeutral)
print (Archive_Data_EmodbEmotionSadness)
print (Archive_Data_AbcAffectAgressiv)
print (Archive_Data_AbcAffectCheerful)
print (Archive_Data_AbcAffectIntoxicatd)
print (Archive_Data_AbcAffectNervous)
print (Archive_Data_AbcAffectNeutral)
print (Archive_Data_AbcAffectTired)
print (Archive_Data_Loi1)
print (Archive_Data_Loi2)
print (Archive_Data_Loi3)"""


"""
data = Get_Data('/Users/paul/Documents/GitHub/ReTiVa3.0/openEAR-0.1.0 Kopie/SmileArchiv/Neue_Session_12:31_28_01_2023.xlsx')
Archive_Data_Time                                = data.get("Archive_Data_Time", "Key not found")
Archive_Data_Aroual                              = data.get("Archive_Data_Arousal", "Key not found")
Archive_Data_Valence                             = data.get("Archive_Data_Valence", "Key not found")
Archive_Data_EmodbEmotionAnger                   = data.get("Archive_Data_EmodbEmotionAnger", "Key not found")
Archive_Data_EmodbEmotionBoredm                  = data.get("Archive_Data_EmodbEmotionBoredom", "Key not found")
Archive_Data_EmodbEmotionDisgust                 = data.get("Archive_Data_EmodbEmotionDisgust", "Key not found")
Archive_Data_EmodbEmotionFear                    = data.get("Archive_Data_EmodbEmotionFear", "Key not found")
Archive_Data_EmodbEmotionHappness                = data.get("Archive_Data_EmodbEmotionHappiness", "Key not found")
Archive_Data_EmodbEmotionNeutral                 = data.get("Archive_Data_EmodbEmotionNeutral", "Key not found")
Archive_Data_EmodbEmotionSadness                 = data.get("Archive_Data_EmodbEmotionSadness", "Key not found")
Archive_Data_AbcAffectAgressiv                   = data.get("Archive_Data_AbcAffectAgressiv", "Key not found")
Archive_Data_AbcAffectCheerful                   = data.get("Archive_Data_AbcAffectCheerfull", "Key not found")
Archive_Data_AbcAffectIntoxicatd                 = data.get("Archive_Data_AbcAffectIntoxicated", "Key not found")
Archive_Data_AbcAffectNervous                    = data.get("Archive_Data_AbcAffectNervous", "Key not found")
Archive_Data_AbcAffectNeutral                    = data.get("Archive_Data_AbcAffectNeutral", "Key not found")
Archive_Data_AbcAffectTired                      = data.get("Archive_Data_AbcAffectTired", "Key not found")
Archive_Data_Loi1                                = data.get("Archive_Data_Loi1", "Key not found")
Archive_Data_Loi2                                = data.get("Archive_Data_Loi2", "Key not found")
Archive_Data_Loi3                                = data.get("Archive_Data_Loi3", "Key not found")
Archive_Soll_DataEmodbEmotionAnger               = data.get("Archive_Soll_DataEmodbEmotionAnger", "Key not found")
Archive_Soll_DataEmodbEmotionBoredom             = data.get("Archive_Soll_DataEmodbEmotionBoredom", "Key not found")
Archive_Soll_DataEmodbEmotionDisgust             = data.get("Archive_Soll_DataEmodbEmotionDisgust", "Key not found")
Archive_Soll_DataEmodbEmotionFear                = data.get("Archive_Soll_DataEmodbEmotionFear", "Key not found")
Archive_Soll_DataEmodbEmotionHappiness           = data.get("Archive_Soll_DataEmodbEmotionHappiness", "Key not found")
Archive_Soll_DataEmodbEmotionNeutral             = data.get("Archive_Soll_DataEmodbEmotionNeutral", "Key not found")
Archive_Soll_DataEmodbEmotionSadness             = data.get("Archive_Soll_DataEmodbEmotionSadness", "Key not found")
Archive_Soll_DataAbcAffectAgressiv               = data.get("Archive_Soll_DataAbcAffectAgressiv", "Key not found")
Archive_Soll_DataAbcAffectCheerfull              = data.get("Archive_Soll_DataAbcAffectCheerfull", "Key not found")
Archive_Soll_DataAbcAffectIntoxicated            = data.get("Archive_Soll_DataAbcAffectIntoxicated", "Key not found")
Archive_Soll_DataAbcAffectNervous                = data.get("Archive_Soll_DataAbcAffectNervous", "Key not found")
Archive_Soll_DataAbcAffectNeutral                = data.get("Archive_Soll_DataAbcAffectNeutral", "Key not found")
Archive_Soll_DataAbcAffectTired                  = data.get("Archive_Soll_DataAbcAffectTired", "Key not found")
Archive_Abs_MW_Data_Arousal                      = data.get("Archive_Abs_MW_Data_Arousal", "Key not found")
Archive_Abs_MW_Data_Valence                      = data.get("Archive_Abs_MW_Data_Valence", "Key not found")
Archive_Abs_MW_Data_EmodbEmotionAnger            = data.get("Archive_Abs_MW_Data_EmodbEmotionAnger", "Key not found")
Archive_Abs_MW_Data_EmodbEmotionBoredom          = data.get("Archive_Abs_MW_Data_EmodbEmotionBoredom", "Key not found")
Archive_Abs_MW_Data_EmodbEmotionDisgust          = data.get("Archive_Abs_MW_Data_EmodbEmotionDisgust", "Key not found")
Archive_Abs_MW_Data_EmodbEmotionFear             = data.get("Archive_Abs_MW_Data_EmodbEmotionFear", "Key not found")
Archive_Abs_MW_Data_EmodbEmotionHappiness        = data.get("Archive_Abs_MW_Data_EmodbEmotionHappiness", "Key not found")
Archive_Abs_MW_Data_EmodbEmotionNeutral          = data.get("Archive_Abs_MW_Data_EmodbEmotionNeutral", "Key not found")
Archive_Abs_MW_Data_EmodbEmotionSadness          = data.get("Archive_Abs_MW_Data_EmodbEmotionSadness", "Key not found")
Archive_Abs_MW_Data_AbcAffectAgressiv            = data.get("Archive_Abs_MW_Data_AbcAffectAgressiv", "Key not found")
Archive_Abs_MW_Data_AbcAffectCheerfull           = data.get("Archive_Abs_MW_Data_AbcAffectCheerfull", "Key not found")
Archive_Abs_MW_Data_AbcAffectIntoxicated         = data.get("Archive_Abs_MW_Data_AbcAffectIntoxicated", "Key not found")
Archive_Abs_MW_Data_AbcAffectNervous             = data.get("Archive_Abs_MW_Data_AbcAffectNervous", "Key not found")
Archive_Abs_MW_Data_AbcAffectNeutral             = data.get("Archive_Abs_MW_Data_AbcAffectNeutral", "Key not found")
Archive_Abs_MW_Data_AbcAffectTired               = data.get("Archive_Abs_MW_Data_AbcAffectTired", "Key not found")
Archive_Abs_MW_Data_Loi1                         = data.get("Archive_Abs_MW_Data_Loi1", "Key not found")
Archive_Abs_MW_Data_Loi2                         = data.get("Archive_Abs_MW_Data_Loi2", "Key not found")
Archive_Abs_MW_Data_Loi3                         = data.get("Archive_Abs_MW_Data_Loi3", "Key not found")
Archive_Score_EmodbEmotions                      = data.get("Archive_Score_EmodbEmotions", "Key not found")
Archive_Score_AbcAffect                          = data.get("Archive_Score_AbcAffect", "Key not found")
Archive_Score_Retiva                             = data.get("Archive_Score_Retiva", "Key not found")
Archive_Abs_MW_Loi_Score                         = data.get("Archive_Abs_MW_Loi_Score", "Key not found")



print(Archive_Data_Time)
print(Archive_Data_Aroual)
print(Archive_Data_Valence)
print(Archive_Data_EmodbEmtionAnger)
print(Archive_Data_EmodbEmotionBoredm)
print(Archive_Data_EmodbEmotionDisgust)
print(Archive_Data_EmodbEmotionFear)
print(Archive_Data_EmodbEmotionHappness)
print(Archive_Data_EmodbEmotionNeutral)
print(Archive_Data_EmodbEmotionSadness)
print(Archive_Data_AbcAffectAgressiv)
print(Archive_Data_AbcAffectCheerful)
print(Archive_Data_AbcAffectIntoxicatd)
print(Archive_Data_AbcAffectNervous)
print(Archive_Data_AbcAffectNeutral)
print(Archive_Data_AbcAffectTired)
print(Archive_Data_Loi1)
print(Archive_Data_Loi2)
print(Archive_Data_Loi3)
print(Archive_Soll_DataEmodbEmotionAnger)
print(Archive_Soll_DataEmodbEmotionBoredom)
print(Archive_Soll_DataEmodbEmotionDisgust)
print(Archive_Soll_DataEmodbEmotionFear)
print(Archive_Soll_DataEmodbEmotionHappiness)
print(Archive_Soll_DataEmodbEmotionNeutral)
print(Archive_Soll_DataEmodbEmotionSadness)
print(Archive_Soll_DataAbcAffectAgressiv)
print(Archive_Soll_DataAbcAffectCheerfull)
print(Archive_Soll_DataAbcAffectIntoxicated)
print(Archive_Soll_DataAbcAffectNervous)
print(Archive_Soll_DataAbcAffectNeutral)
print(Archive_Soll_DataAbcAffectTired)
print(Archive_Abs_MW_Data_Arousal)
print(Archive_Abs_MW_Data_Valence)
print(Archive_Abs_MW_Data_EmodbEmotionAnger)
print(Archive_Abs_MW_Data_EmodbEmotionBoredom)
print(Archive_Abs_MW_Data_EmodbEmotionDisgust)
print(Archive_Abs_MW_Data_EmodbEmotionFear)
print(Archive_Abs_MW_Data_EmodbEmotionHappiness)
print(Archive_Abs_MW_Data_EmodbEmotionNeutral)
print(Archive_Abs_MW_Data_EmodbEmotionSadness)
print(Archive_Abs_MW_Data_AbcAffectAgressiv)
print(Archive_Abs_MW_Data_AbcAffectCheerfull)
print(Archive_Abs_MW_Data_AbcAffectIntoxicated)
print(Archive_Abs_MW_Data_AbcAffectNervous)
print(Archive_Abs_MW_Data_AbcAffectNeutral)
print(Archive_Abs_MW_Data_AbcAffectTired)
print(Archive_Abs_MW_Data_Loi1)
print(Archive_Abs_MW_Data_Loi2)
print(Archive_Abs_MW_Data_Loi3)
print(Archive_Score_EmodbEmotions)
print(Archive_Score_AbcAffect)
print(Archive_Score_Retiva)
print(Archive_Abs_MW_Loi_Score)
"""



    

















































